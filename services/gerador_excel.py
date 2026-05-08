from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from copy import copy
from openpyxl.worksheet.page import PageMargins

def copiar_estilo(ws, linha_origem, linha_destino):
    for col in range(1, 8):
        origem = ws.cell(row=linha_origem, column=col)
        destino = ws.cell(row=linha_destino, column=col)

        if origem.has_style:
            destino._style = copy(origem._style)

        destino.number_format = origem.number_format
        destino.alignment = copy(origem.alignment)
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)

def gerar_relatorio(dados_mes, ano, mes, totais, data_final, caminho_template, nome_arquivo):

    dados_mes["tipo"] = dados_mes["tipo"].str.strip().str.lower()

    wb = load_workbook(caminho_template)
    ws = wb.active

    linhas_base = 20
    quantidade = len(dados_mes)
    extras = 0

    if quantidade > linhas_base:
        extras = quantidade - linhas_base
        ws.insert_rows(28, extras)

        for i in range(extras):
            copiar_estilo(ws, 8, 28 + i)

    elif quantidade < linhas_base:
        remover = linhas_base - quantidade
        ws.delete_rows(quantidade + 8, remover)

    linha = 8 # Início dos dados

    print("TIPOS NO GERADOR:")
    print(dados_mes["tipo"].apply(repr).unique())

    # ===== DADOS =====
    for _, row in dados_mes.iterrows():

        # DATA
        cell_data = ws.cell(row=linha, column=1, value=row["data"])
        cell_data.number_format = 'DD/MM'
        cell_data.alignment = Alignment(horizontal='center')

        # HISTÓRICO
        ws.cell(row=linha, column=2, value=row["historico"])

        # ENTRADAS
        tipo = row["tipo"].strip().lower()

        if tipo == "entrada":
            ws.cell(row=linha, column=4, value=row["valor"]).number_format = '#,##0.00'
        else:
            ws.cell(row=linha, column=4, value="")

        # SAÍDAS
        if row["tipo"] == "saida":
            ws.cell(row=linha, column=5, value=row["valor"]).number_format = '#,##0.00'
        else:
            ws.cell(row=linha, column=5, value="")

        # DOADOR/BENEF
        if row["tipo"] == "doador":
            ws.cell(row=linha, column=6, value=row["valor"]).number_format = '#,##0.00'
        else:
            ws.cell(row=linha, column=6, value="")

        linha += 1

        print(f"Tipo linha: ->{row['tipo']}<-")

    # ===== TOTAIS =====
    total_entradas, total_saidas, saldo = totais

    linha_total = linha + 1
    
    # TOTAIS ENTRADAS
    ws.merge_cells(start_row=linha_total, start_column=2, end_row=linha_total, end_column=3)
    label_entrada = ws.cell(row=linha_total, column=2, value="TOTAL DAS ENTRADAS")
    label_entrada.font = Font(bold=True)
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.cell(row=linha_total, column=4, value=total_entradas).number_format = 'R$ #,##0.00'

    # TOTAL DAS SAÍDAS
    ws.merge_cells(start_row=linha_total+1, start_column=2, end_row=linha_total+1, end_column=3)
    label_saida = ws.cell(row=linha_total+1, column=2, value="TOTAL DAS SAÍDAS")
    label_saida.font = Font(bold=True)
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.cell(row=linha_total+1, column=4, value=total_saidas).number_format = 'R$ #,##0.00'

    # SALDO
    ws.merge_cells(start_row=linha_total+2, start_column=2, end_row=linha_total+2, end_column=3)
    label_saldo = ws.cell(row=linha_total+2, column=2, value="SALDO TOTAL")
    label_saldo.font = Font(bold=True)
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.cell(row=linha_total+2, column=4, value=saldo).number_format = 'R$ #,##0.00'

    # ===== RODAPÉ =====
    linha_base = linha_total + 4

    # LOCAL
    label_local = ws.cell(row=linha_base, column=1, value="LOCAL")
    label_local.alignment = Alignment(horizontal='center')
    label_local.font = Font(bold=True)
    
    ws.merge_cells(start_row=linha_base, start_column=2, end_row=linha_base, end_column=4)
    valor_local = ws.cell(row=linha_base, column=2, value=" ")
    valor_local.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=linha_base, start_column=5, end_row=linha_base+1, end_column=5)
    ws.merge_cells(start_row=linha_base, start_column=6, end_row=linha_base+1, end_column=6)

    # DATA
    label_data = ws.cell(row=linha_base+1, column=1, value="DATA")
    label_data.alignment = Alignment(horizontal='center')
    label_data.font = Font(bold=True)

    ws.merge_cells(start_row=linha_base+1, start_column=2, end_row=linha_base+1, end_column=4)
    valor_data = ws.cell(row=linha_base+1, column=2, value=data_final)
    valor_data.number_format = 'DD/MM/YYYY'
    valor_data.alignment = Alignment(horizontal='center', vertical='center')
    valor_data.font = Font(bold=True)

    ws.merge_cells(start_row=linha_base+1, start_column=5, end_row=linha_base+2, end_column=5)
    ws.merge_cells(start_row=linha_base+1, start_column=6, end_row=linha_base+2, end_column=6)

    # ASSINATURA
    label_ass = ws.cell(row=linha_base+2, column=1, value="ASSINATURA\nDO DIRETOR")
    ws.merge_cells(start_row=linha_base+2, start_column=1, end_row=linha_base+3, end_column=1)
    label_ass.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions["A"].width = 18
    label_ass.font = Font(bold=True)

    ws.merge_cells(start_row=linha_base+2, start_column=2, end_row=linha_base+3, end_column=4)
    valor_ass = ws.cell(row=linha_base+2, column=2, value=" ")
    ws.merge_cells(start_row=linha_base+2, start_column=5, end_row=linha_base+3, end_column=5)
    ws.merge_cells(start_row=linha_base+2, start_column=6, end_row=linha_base+3, end_column=6)

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"

    ws.sheet_view.zoomScale = 100

    ws.page_margins = PageMargins(
        left=0.3,
        right=0.3,
        top=0.5,
        bottom=0.5
    )

    ultima_linha = linha_base + 3
    ws.print_area = f"A1:F{ultima_linha}"

    wb.save(nome_arquivo)